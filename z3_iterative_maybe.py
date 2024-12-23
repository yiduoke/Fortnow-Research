from z3 import *
import random
from openpyxl import Workbook
from openpyxl.styles import PatternFill


# Initialize solver
solver = Solver()

# Simulated parameters
num_items = 10  # Number of items
random.seed(42)

items = [f"item_{i}" for i in range(num_items)]
u_A1 = {item: random.randint(1, 20) for item in items} # Agent 1's utility for his own items
u_A2 = {item: random.randint(1, 20) for item in items} # Agent 2's utility for his own items

# Print utility table
print("\nUtility Table:")
print(f"{'Item':<10}{'Utility A1':<15}{'Utility A2':<15}")
for item in items:
    print(f"{item:<10}{u_A1[item]:<15}{u_A2[item]:<15}")

# Turing machine outputs
tm_A = sorted(items, key=lambda item: u_A1[item], reverse=True)
tm_B = sorted(items, key=lambda item: u_A2[item], reverse=True)

# Merge outputs from TM A and TM B
revealed_items = []
for i in range(max(len(tm_A), len(tm_B))):
    if i < len(tm_A):
        revealed_items.append(tm_A[i])
    if i < len(tm_B):
        revealed_items.append(tm_B[i])

# Variables for allocation
x = {item: Bool(f"x_{item}") for item in items}

# EFX constraints
def add_efx_constraints(solver, revealed, alloc, u_A1, u_A2):
  # Utilities for agent's own bundle
  U_A1 = Sum([If(alloc[item], u_A1[item], 0) for item in revealed])
  U_A2 = Sum([If(Not(alloc[item]), u_A2[item], 0) for item in revealed])

  # Utilities for the other agent's bundle:
  U_A1_2 = Sum([If(Not(alloc[item]), u_A1[item], 0) for item in revealed])
  U_A2_1 = Sum([If(alloc[item], u_A2[item], 0) for item in revealed])

  # Add EFX constraints for revealed items
  for item in revealed:
    solver.add(U_A1 >= U_A1_2 - If(Not(alloc[item]), u_A1[item], U_A1_2))
    solver.add(U_A2 >= U_A2_1 - If(alloc[item], u_A2[item], U_A2_1))

# Maintain a set of solutions
valid_allocations = []

# Process revealed items
for step in range(len(revealed_items)):
  current_pool = revealed_items[:step + 1] # Current pool of revealed items by the TMs
  solver.push()

  # Add EFX constraints
  add_efx_constraints(solver, current_pool, x, u_A1, u_A2)

  # Find all valid solutions for the current pool
  current_allocations = []
  while solver.check() == sat:

    model = solver.model()
    solution = {item: model.evaluate(x[item]) for item in current_pool}
    allocation = {item: model[x[item]] for item in items}

    current_allocations.append(allocation)

    # Adds a constraint to block the current solution so the solver can find different solutions
    solver.add(Not(And([x[item] == solution[item] for item in current_pool])))
    

  # Filter solutions: Keep only subsets of the next step's allocations consistent with any previous ones
  if valid_allocations:
    valid_allocations = [
      alloc for alloc in current_allocations
      if any(all(alloc[item] == valid[item] for item in valid) for valid in valid_allocations)
    ]
  else:
    valid_allocations = current_allocations

  solver.pop()
  


def generate_excel(valid_allocations, revealed_items, u_A1, u_A2, filename="allocations.xlsx"):
  # Create a new Excel workbook
  workbook = Workbook()

  # Define fill colors for highlighting
  color_fill = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")  # Light orange

  # Loop through each valid allocation
  for idx, allocation in enumerate(valid_allocations, 1):
    # Create a new sheet for the allocation
    sheet = workbook.create_sheet(title=f"Allocation {idx}")

    # Add headers
    sheet.append([
      "Item", 
      "Player 1 Utility", "Player 2 Utility", 
      "Player 1 Cumulative Utility", "Player 2 Cumulative Utility",
      "Player 1 Utility on Player 2's Bundle", "Player 2 Utility on Player 1's Bundle"
    ])

    # Initialize cumulative utilities
    player_1_cumulative = 0
    player_2_cumulative = 0
    player_1_on_player_2_bundle = 0
    player_2_on_player_1_bundle = 0

    # Populate rows
    for item in revealed_items:
      # Player utilities for the current item
      p1_utility = u_A1[item]
      p2_utility = u_A2[item]

      # Update cumulative utilities based on allocation
      if allocation[item]:  # Assigned to Player 1
        player_1_cumulative += p1_utility
        player_2_on_player_1_bundle += p2_utility
      else:  # Assigned to Player 2
        player_2_cumulative += p2_utility
        player_1_on_player_2_bundle += p1_utility

      # Add the row
      row = [
        item, 
        p1_utility, p2_utility, 
        player_1_cumulative, player_2_cumulative,
        player_1_on_player_2_bundle, player_2_on_player_1_bundle
      ]
      sheet.append(row)

      # Get the row index (starts at 2 because of the header)
      row_idx = sheet.max_row

      # Highlight the cell for the assigned player
      if allocation[item]:  # Assigned to Player 1
          sheet.cell(row=row_idx, column=2).fill = color_fill
      else:  # Assigned to Player 2
          sheet.cell(row=row_idx, column=3).fill = color_fill

    # Adjust column widths for readability
    for col in sheet.columns:
      sheet.column_dimensions[col[0].column_letter].width = 20  # Adjust as needed

  # Remove the default sheet created by openpyxl
  if "Sheet" in workbook.sheetnames:
    workbook.remove(workbook["Sheet"])

  # Save the workbook
  workbook.save(filename)
  print(f"Excel file '{filename}' has been created successfully.")

# Example usage
# valid_allocations: List of dictionaries, each representing a final valid allocation
# revealed_items: List of items in the order they were revealed
# u_A1, u_A2: Utility dictionaries for Player 1 and Player 2
generate_excel(valid_allocations, items, u_A1, u_A2)

